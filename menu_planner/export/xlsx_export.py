"""
export/xlsx_export.py — Export weekly menu to Excel with exact formatting.
"""
from __future__ import annotations
from pathlib import Path
from datetime import date, timedelta

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

from config import DAYS, SLOTS, SLOT_IDS, SLOT_DISPLAY, SLOT_SECTION, SECTIONS


# Section order for rows (in display order)
SECTION_ORDER = [
    ("LUNCH",      "Lunch"),
    ("SNACKS",     "Morning & Evening Snacks"),
    ("FRUIT",      "Fruit / Healthy (Diet) Lunch"),
    ("ADDITIONAL", "Additional"),
]

SECTION_FILLS = {
    "LUNCH":      "D0D0D0",
    "SNACKS":     "E8E8E8",
    "FRUIT":      "D8D8D8",
    "ADDITIONAL": "C8C8C8",
}

FRUIT_LABEL = "FRUIT / HEALTHY (DIET) LUNCH"


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, size=10, color="1A1A1A", name="Calibri") -> Font:
    return Font(name=name, bold=bold, size=size, color=color)


def _align(h="center", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _border(style="thin") -> Border:
    s = Side(style=style, color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


def _style_range(ws, cell_range, font=None, fill=None, border=None, alignment=None):
    """Style all cells in a range (useful for merged ranges in openpyxl)."""
    for row in ws[cell_range]:
        for cell in row:
            if font is not None: cell.font = font
            if fill is not None: cell.fill = fill
            if border is not None: cell.border = border
            if alignment is not None: cell.alignment = alignment


def export_xlsx(
    week_data: dict,
    week_start: date,
    output_path: str | Path,
    company_name: str = "Chathradhari Caterers",
    canteen_name: str = "Oyster",
) -> Path:
    """
    Export a week's menu to Excel.

    week_data: {DAY: {slot_id: dish_name}}
    week_start: date of Monday
    output_path: where to save the .xlsx
    """
    if openpyxl is None:
        raise ImportError("openpyxl is required: pip install openpyxl")

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Weekly Menu"

    # Column widths: A,B=5, C=20, D-H=18
    col_widths = [5, 5, 20, 18, 18, 18, 18, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Build day->date map
    day_dates = {day: week_start + timedelta(days=i) for i, day in enumerate(DAYS)}

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    ws["A1"] = "WEEKLY LUNCH / SNACK MENU"
    _style_range(
        ws, "A1:H1",
        font=_font(bold=True, size=14, color="FFFFFF"),
        fill=_fill("7A7A7A"),
        alignment=_align()
    )
    ws.merge_cells("A1:H1")
    ws.row_dimensions[1].height = 35

    # ── Row 2: empty ─────────────────────────────────────────────────────────
    ws.row_dimensions[2].height = 8

    # ── Row 3: notice text ────────────────────────────────────────────────────
    ws["A3"] = (
        f"Note: Menu is subject to change. {company_name} — {canteen_name} Canteen. "
        "All items prepared fresh daily."
    )
    _style_range(
        ws, "A3:H3",
        font=_font(size=9, color="555555"),
        alignment=_align(h="left")
    )
    ws.merge_cells("A3:H3")
    ws.row_dimensions[3].height = 16

    # ── Row 4: empty ─────────────────────────────────────────────────────────
    ws.row_dimensions[4].height = 8

    # ── Row 5: Column headers ─────────────────────────────────────────────────
    headers = ["", "", "LIST OF ITEMS"] + [d for d in DAYS]
    for col, val in enumerate(headers, 1):
        c = ws.cell(row=5, column=col, value=val)
        c.font = _font(bold=True, size=10, color="FFFFFF")
        c.fill = _fill("555555")
        c.alignment = _align()
        c.border = _border()
    ws.row_dimensions[5].height = 20

    # ── Row 6: Dates row ──────────────────────────────────────────────────────
    date_vals = ["", "", "DATE"] + [
        day_dates[d].strftime("%d-%b-%y") for d in DAYS
    ]
    for col, val in enumerate(date_vals, 1):
        c = ws.cell(row=6, column=col, value=val)
        c.font = _font(bold=True, size=10, color="FFFFFF")
        c.fill = _fill("3E2007")
        c.alignment = _align()
        c.border = _border()
    ws.row_dimensions[6].height = 22

    # ── Data rows ─────────────────────────────────────────────────────────────
    current_row = 7
    row_num_col = 1   # col A: sequential number within section
    slot_counter = {sec: 0 for sec in SECTIONS}

    # Build slot list grouped by section
    section_slots: dict[str, list] = {sec: [] for sec in SECTIONS}
    for slot_id, display, section, ppt_sec, is_fixed in SLOTS:
        section_slots[section].append((slot_id, display))

    fruit_header_written = False

    for sec_key, sec_label in SECTION_ORDER:
        sec_fill_hex = SECTION_FILLS.get(sec_key, "E0E0E0")

        # Section header row
        if sec_key == "FRUIT":
            # Special merged left label for fruit section (written alongside rows)
            fruit_start_row = current_row
            fruit_header_written = True
        else:
            ws.cell(row=current_row, column=1, value=sec_label.upper())
            _style_range(
                ws, f"A{current_row}:H{current_row}",
                font=_font(bold=True, size=11, color="FFFFFF"),
                fill=_fill("444444"),
                border=_border(),
                alignment=_align()
            )
            ws.merge_cells(f"A{current_row}:H{current_row}")
            ws.row_dimensions[current_row].height = 20
            current_row += 1

        slots_in_sec = section_slots[sec_key]
        fruit_data_start = current_row

        for idx, (slot_id, display) in enumerate(slots_in_sec):
            row_fill = _fill("EBF5FB") if idx % 2 == 0 else _fill("D6EAF8")
            if sec_key != "FRUIT":
                # Col A: row number
                c = ws.cell(row=current_row, column=1, value=idx + 1)
                c.font = _font(size=9)
                c.alignment = _align()
                c.fill = row_fill
                c.border = _border()
                # Col B: blank
                c = ws.cell(row=current_row, column=2, value="")
                c.fill = row_fill
                c.border = _border()
                # Col C: slot name
                c = ws.cell(row=current_row, column=3, value=display)
                c.font = _font(bold=True, size=10)
                c.alignment = _align(h="left")
                c.fill = row_fill
                c.border = _border()
            else:
                # FRUIT section: cols A-B merged as label on first row
                if idx == 0:
                    # Will merge later
                    pass
                c = ws.cell(row=current_row, column=3, value=display)
                c.font = _font(bold=True, size=10)
                c.alignment = _align(h="left")
                c.fill = row_fill
                c.border = _border()

            # Cols D-H: day values
            for d_idx, day in enumerate(DAYS):
                dish = week_data.get(day, {}).get(slot_id, "")
                col = 4 + d_idx
                c = ws.cell(row=current_row, column=col, value=dish)
                c.font = _font(bold=bool(dish), size=10)
                c.alignment = _align(h="center", wrap=True)
                c.fill = row_fill
                c.border = _border()

            ws.row_dimensions[current_row].height = 20
            current_row += 1

        # For FRUIT section: merge A:B for the section label
        if sec_key == "FRUIT" and slots_in_sec:
            fruit_end_row = current_row - 1
            ws.cell(row=fruit_data_start, column=1, value=FRUIT_LABEL)
            _style_range(
                ws, f"A{fruit_data_start}:B{fruit_end_row}",
                font=_font(bold=True, size=9, color="FFFFFF"),
                fill=_fill("444444"),
                border=_border(),
                alignment=_align(h="center", wrap=True)
            )
            ws.merge_cells(f"A{fruit_data_start}:B{fruit_end_row}")

        # Gap row
        ws.row_dimensions[current_row].height = 6
        current_row += 1

    # ── Freeze top rows ────────────────────────────────────────────────────────
    ws.freeze_panes = "D7"

    wb.save(output_path)
    return output_path
